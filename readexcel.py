import openpyxl
import pprint


def change_to_number(listA):
    value = 0
    for x in listA:
        x= int(x.replace('s',''))
        value += (((x - 4) / (2 * 4)) + 0.5)
    return value / (len(listA))

def  normal_decision_matrix(listA,listB):
    for y in range(len(listA[0])):
        zx=0
        for z in listB:
            sumx=0
            for x in range(z):
                sumx+=listA[zx+x][y]
            #print(sumx)
            for x in range(z):
                listA[zx+x][y]=listA[zx+x][y]/sumx
            #print(lis
            zx+=z
    return listA



wb = openpyxl.load_workbook("D:\\work\\Python\\supply chain\\anp.xlsx")
sh = wb.active
data = []
data_w=[]
data_w.append(sh.cell(row=25, column=2).value)
data_w.append(sh.cell(row=26, column=2).value)
data_w.append(sh.cell(row=27, column=2).value)
data_w.append(sh.cell(row=28, column=2).value)
weight= []
for x in range(15):
    data.append(list([]))
    for y in range(15):
        c3 = sh.cell(row=2 + x, column=2 + y)
        value = c3.value.split(',')
        value =change_to_number(value)
        data[x].append(value)

for x in range(4):
    weight.append(list([]))
    for y in range(4):
        c3 = sh.cell(row=20 + x, column=2 + y)
        value = float(c3.value)
        weight[x].append(value)
print(weight)
weight=normal_decision_matrix(weight,[4])

wb.close()

wb = openpyxl.Workbook()
sh = wb.active
print(data_w)
print(data)
data=normal_decision_matrix(data,data_w)


for x in range(15):
    for y in range(15):
        c3 = sh.cell(row=2 + x, column=2 + y)
        c3.value=data[x][y]

for x in range(4):
    for y in range(4):
        c3 = sh.cell(row=20 + x, column=2 + y)
        c3.value=weight[x][y]
wb.save("output.xlsx")
wb.close()
#pprint.pprint(data)
