import openpyxl
import pprint
import numpy as np
from fractions import Fraction
from supermatrix import calculate_limit_supermat


def change_to_number(listA):
    value = 0
    for x in listA:
        x = Fraction(x.replace('s', ''))
        value += (((x - 4) / (2 * 4)) + 0.5)
        #   value+= x
    return value / (len(listA))


def normalize_decision_matrix(listA):
    for y in range(len(listA[0])):
        sumx = 0
        for x in range(len(listA)):
            sumx += listA[x][y]
        if sumx == 0:
            continue
        for x in range(len(listA)):
            listA[x][y] = listA[x][y] / sumx
    return (listA)


def average_matris(listA):
    b = []

    for x in listA:
        b.append(sum(x))
    sumx = sum(b)
    # print(listA,"bbbbb",b,sumx)
    if sumx == 0:
        pass
    else:
        for x in range(len(b)):
            b[x] /= sumx
    return b


def opendata():
    wb = openpyxl.load_workbook("D:\\work\\Python\\supply chain\\input.xlsx")
    sh = wb.active
    data = []
    c3 = sh.cell(row=1, column=1)
    Nocolmun = int(c3.value)
    c3 = sh.cell(row=1, column=2)
    NoRow = int(c3.value)
    for x in range(Nocolmun * NoRow):
        data.append(list([]))
        row = int(sh.cell(row=2 + x, column=1).value)
        colmun = int(sh.cell(row=2 + x, column=2).value)
        numbermatris = int(sh.cell(row=2 + x, column=3).value)
        temp_colmun = 3
        for m in range(numbermatris):
            if row == 2:
                temp_colmun += 1
                c3 = sh.cell(row=2 + x, column=temp_colmun)
                if (c3.value == 0):
                    data[x].append(np.zeros((row, colmun)))
                else:
                    f = c3.value.split(",")
                    f = change_to_number(f)
                    t = np.identity(row) * 1 / 2
                    t[0][1] = 1 - f
                    t[1][0] = f
                    data[x].append(t)
            elif row == 3:
                temp_colmun += 1
                c3 = sh.cell(row=2 + x, column=temp_colmun)
                c4 = sh.cell(row=2 + x, column=temp_colmun + 1)
                c5 = sh.cell(row=2 + x, column=temp_colmun + 2)
                if (c3.value == 0):
                    data[x].append(np.zeros((row, colmun)))
                else:
                    f = change_to_number(c3.value.split(","))
                    f1 = change_to_number(c4.value.split(","))
                    f2 = change_to_number(c5.value.split(","))
                    t = np.identity(row) * 0.5
                    t[0][1] = 1 - f
                    t[0][2] = 1 - f1
                    t[1][2] = 1 - f2
                    t[1][0] = f
                    t[2][0] = f1
                    t[2][1] = f2
                    temp_colmun += 2
                    data[x].append(t)
            elif row == 4:
                temp_colmun += 1
                c3 = sh.cell(row=2 + x, column=temp_colmun)
                c4 = sh.cell(row=2 + x, column=temp_colmun + 1)
                c5 = sh.cell(row=2 + x, column=temp_colmun + 2)
                c6 = sh.cell(row=2 + x, column=temp_colmun + 3)
                c7 = sh.cell(row=2 + x, column=temp_colmun + 4)
                c8 = sh.cell(row=2 + x, column=temp_colmun + 5)
                if (c3.value == 0):
                    data[x].append(np.zeros((row, colmun)))
                else:
                    f = change_to_number(c3.value.split(","))
                    f1 = change_to_number(c4.value.split(","))
                    f2 = change_to_number(c5.value.split(","))
                    f3 = change_to_number(c6.value.split(","))
                    f4 = change_to_number(c7.value.split(","))
                    f5 = change_to_number(c8.value.split(","))
                    t = np.identity(row) * 0.5
                    t[0][1] = 1 - f
                    t[0][2] = 1 - f1
                    t[0][3] = 1 - f2
                    t[1][2] = 1 - f3
                    t[1][3] = 1 - f4
                    t[2][3] = 1 - f5
                    t[1][0] = f
                    t[2][0] = f1
                    t[3][0] = f2
                    t[2][1] = f3
                    t[3][1] = f4
                    t[3][2] = f5
                    temp_colmun += 5
                    data[x].append(t)
    # pprint.pprint(data[9])
    # for y in range(0, row):
    #     data[x][m].append([])
    #     for z in range(colmun):
    #         temp_colmun += 1
    #         c3 = sh.cell(row=2 + x, column=temp_colmun)
    #         if (c3.value == 0):
    #             data[x][m][y].append(0)
    #         else:
    #             value = c3.value.split(',')
    #             data[x][m][y].append(change_to_number(value))
    weight = []
    # print(data[9])
    for x in range(NoRow):
        weight.append(list([]))
        for y in range(NoRow):
            c3 = sh.cell(row=(Nocolmun * NoRow) + 2 + x, column=1 + y)
            # print(c3.value)
            value = float(c3.value)
            weight[x].append(value)
    # print("weight2", weight)
    weight = normalize_decision_matrix(weight)
    # print("weight2", weight)
    NoCluster = []
    for y in range(NoRow):
        c3 = sh.cell(row=1, column=3 + y)
        value = int(c3.value)
        NoCluster.append(value)
    for z in range(len(data)):
        for t in range(len(data[z])):
            data[z][t] = normalize_decision_matrix(data[z][t])
    # print(data[0])
    for z in range(len(data)):
        # print(z,type(data[z]))
        for t in range(len(data[z])):
            data[z][t] = average_matris(data[z][t])
    # print(data)
    # pprint.pprint(data[9])
    # print(data[4])
    for x in range(0, NoRow * NoRow, NoRow + 1):
        temp = len(data[x][0])
        if temp == 2:
            if data[x][0][0] != 0 and data[x][0][1] != 0:
                t = np.array([[0, data[x][1][0], data[x][2][0]],
                              [data[x][0][0], 0, data[x][2][1]],
                              [data[x][0][1], data[x][1][1], 0]])
                # print("sssssssssssssssssss",t)
                data[x] = t
        else:
            # print(x,data[x])
            t = np.array([[0, data[x][1][0], data[x][2][0], data[x][3][0]],
                          [data[x][0][0], 0, data[x][2][1], data[x][3][1]],
                          [data[x][0][1], data[x][1][1], 0, data[x][3][2]],
                          [data[x][0][2], data[x][1][2], data[x][2][2], 0]])
            data[x] = t
    finaldata = []
    temp = -1
    # pprint.pprint(data[9])
    for x in range(NoRow):
        finaldata.append([])
        for y in range(NoRow):
            temp += 1
            if x == y:
                finaldata[x].append(np.array(data[temp]) * weight[x][y])
            else:
                finaldata[x].append(np.array(data[temp]).transpose() * weight[x][y])

    for x in range(NoRow):
        # print("---------------------------")
        # pprint.pprint(finaldata[x])
        finaldata[x] = np.concatenate(finaldata[x], axis=1)
    print("++++++++++++++++++++++++++")
    # pprint.pprint(finaldata[3])
    # for x in range(4):
    #     print(len(finaldata[x][0]))
    # pprint.pprint(finaldata[-2])
    # pprint.pprint(weight)
    finaldata = np.concatenate(finaldata, axis=0)
    answer = calculate_limit_supermat(finaldata, 0.00000001)
    print("--------------------------")
    # print("ans is ",answer[0]/(answer[0]+answer[1]+answer[2]+answer[3]))
    print("weight is all in")
    pprint.pprint(answer)
    temp = 0
    anscluster = []
    for x in NoCluster:
        sumx = sum(answer[temp:(temp + x)])
        anscluster.append([])
        for y in range(x):
            anscluster[-1].append(answer[y + temp] / sumx)
        temp += x
    print("weight is cluster in")
    pprint.pprint(anscluster)
    print("++++++++++++++")
    sumx = sum(answer[0:11])
    enablears = []
    for x in range(12):
        enablears.append(answer[x] / sumx)

    print("weight is enablears in")
    pprint.pprint(enablears)

    # print(finaldata)
    # print(type(finaldata))
    # print(len(finaldata))


if __name__ == "__main__":
    opendata()
